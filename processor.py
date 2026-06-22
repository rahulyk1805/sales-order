"""
processor.py
────────────
Core business logic: reads uploaded sales-order data and produces a
fully-formatted, multi-sheet Pending Orders workbook in memory.

TWO INPUT MODES — auto-detected by the `process()` entry point:

  Mode A  —  Two raw JSSOReport exports  (the original 2-file flow)
             Both files must have a sheet named 'JSSOReport'.
             The app derives brand, depot, SKU, qty-in-kg etc. from the
             raw columns just like the original desktop script.

  Mode B  —  Single pre-built Raw Data file
             One file with a sheet named 'Raw Data' whose columns already
             match the Raw Data output sheet (Brand / Depot / SKU Display /
             Order Type / Pending (kg or ltr) etc. are pre-computed).
             The app maps those columns to the internal format and skips
             all the derivation logic.

Both modes feed the same `_build_workbook()` function and produce the same
4-sheet output: Production Planning, Order Queue (FIFO), External Orders,
Raw Data.
"""

import io
from datetime import datetime
from collections import defaultdict
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────
# Business rules / lookup tables
# ─────────────────────────────────────────────────────────────────────────
BRAND_STRIP = {'GAIN', 'NUTRIZIN', 'SUPER', 'ZIDDHI', 'ZIDDI', 'INSTRACAL',
               'INSTABOR', 'INSTASUL', 'KARNALZD', 'GAINMEERUT', 'MP'}
WDG_BRANDS = {'Hariyali DF', 'Super Hariyali', 'Profit DF', 'Sulphur Powder'}
MAX_RATE = 200

DEPOT_ORDER = ['Hisar', 'Karnal', 'Bathinda', 'Jaipur', 'Ahmedabad',
               'Pune', 'Akola', 'Indore', 'Raipur', 'Meerut', 'Lucknow']

WDG_BRAND_ORDER   = ['Hariyali DF', 'Super Hariyali', 'Profit DF']
BRAND_BRAND_ORDER = ['Gain', 'InstaSul', 'Ziddi DF', 'InstaBor 150', 'InstaCal 160', 'NutriZin']

# Columns that uniquely identify a pre-built Raw Data sheet (Mode B)
RAW_DATA_REQUIRED_COLS = {
    'Order Type', 'Brand', 'Depot', 'Customer Name', 'SO / PO No.',
    'Order Date', 'Sched. Date', 'Product Name', 'SKU Display',
    'Pending (kg or ltr)', 'Unit',
}

SKU_CLEAN = {
    # Hariyali DF
    'Haryali DF 80% WDG - 30kg ( 5x 6 Pouches)':          'Hariyali DF 30 kg (5×6 Pouches)',
    'Haryali DF 80% WDG - 30kg Bags (3x10)':               'Hariyali DF 30 kg Bags (3×10)',
    'Haryali DF 80% WDG 30 Kg Drum (3x10)':                'Hariyali DF 30 kg Drum (3×10)',
    'Haryali DF 80% WDG -15 Kg Loose Bag':                 'Hariyali DF 15 kg Loose Bag',
    'Haryali Df 80% 30 KG Bags Loose':                     'Hariyali DF 30 kg Loose Bags',
    'Haryali Df 80% WDG - 25 Kgs':                         'Hariyali DF 25 kg',
    'Haryali Df 80% WDG - 25 Kgs Loose Drums':             'Hariyali DF 25 kg Loose Drums',
    'Haryali Df-80%-Wdg-15 Kgs Buckets':                   'Hariyali DF 15 kg Bucket',
    'Haryali-Df-80%-Wdg-15 Kgs Buckets':                   'Hariyali DF 15 kg Bucket',
    'Haryali Df-80%Wdg-50kgs':                             'Hariyali DF 50 kg',
    'Haryali-Df-80%Wdg-50kgs':                             'Hariyali DF 50 kg',
    'Haryali Df-80%WDG- 1 Kg x 20kgs Outer Carton(1x20)': 'Hariyali DF 1 kg × 20 (Carton)',
    'Haryali DF - 80% WDG':                                'Hariyali DF (Bulk)',
    # Super Hariyali
    'Haryali Super 90% WG  - 10 Kgs Bucket':               'Super Hariyali 10 kg Bucket',
    'Haryali Super 90% WG  - 15 Kgs Bucket':               'Super Hariyali 15 kg Bucket',
    'Haryali Super 90% WG 15 KG Loose Bags':               'Super Hariyali 15 kg Loose Bags',
    'Haryali Super 90% WG 30 KG Bags (3x10)':              'Super Hariyali 30 kg Bags (3×10)',
    'Haryali Super 90% WG 30 KG Drums 3x10':               'Super Hariyali 30 kg Drums (3×10)',
    'Haryali Super 90% WG 30 KG Loose Bags':               'Super Hariyali 30 kg Loose Bags',
    'Haryali Super 90% WG 30 Kgs Loose Drum':              'Super Hariyali 30 kg Loose Drum',
    # Profit DF
    'Profit 80 % Wdg 15 Kg Loose Bags':                    'Profit DF 15 kg Loose Bags',
    'Profit 80 % Wdg 30 Kg Loose Bags':                    'Profit DF 30 kg Loose Bags',
    'Profit 80 % Wdg 30 Kg Loose Drums':                   'Profit DF 30 kg Loose Drums',
    'Profit DF 80% WDG (1 Kg x 20)':                       'Profit DF 1 kg × 20',
    'Profit Df 80%- 15 Kg Bucket':                         'Profit DF 15 kg Bucket',
    'Profit Df 80%- 30 Kg Bags (3x10)':                    'Profit DF 30 kg Bags (3×10)',
    # Gain
    'GAIN 500 ML':  'Gain 500 ml',
    'GAIN 1 LTR':   'Gain 1 Ltr',
    # Ziddi DF
    'Ziddi DF - 250gm': 'Ziddi DF 250 gm',
    'Ziddi DF - 500gm': 'Ziddi DF 500 gm',
    'Ziddi DF - 1 Kg':  'Ziddi DF 1 kg',
    # NutriZin
    'Nutrizin 700- 100ML': 'NutriZin 700 100 ml',
    'Nutrizin 700- 250ML': 'NutriZin 700 250 ml',
    'Nutrizin 700- 500ML': 'NutriZin 700 500 ml',
    'Nutrizin 700- 1 Ltr': 'NutriZin 700 1 Ltr',
    # InstaBor
    'Insta Bor 150 (100ML)': 'InstaBor 150 100 ml',
    'Insta Bor 150 (250ML)': 'InstaBor 150 250 ml',
    'Insta Bor 150 (500ML)': 'InstaBor 150 500 ml',
    'Insta Bor 150 (1 Ltr)': 'InstaBor 150 1 Ltr',
    # InstaCal
    'Insta Cal 160 (250ML)': 'InstaCal 160 250 ml',
    'Insta Cal 160 (500ML)': 'InstaCal 160 500 ml',
    'Insta Cal 160 (1 Ltr)': 'InstaCal 160 1 Ltr',
    'Insta Cal 160 (5 Ltr)': 'InstaCal 160 5 Ltr',
    # InstaSul
    'Insta Sul 800 (1 Ltr)':   'InstaSul 800 1 Ltr',
    'INSTASUL 800 (500 ML)':   'InstaSul 800 500 ml',
    'INSTASUL 800 (5 LTR)':    'InstaSul 800 5 Ltr',
}


# ─────────────────────────────────────────────────────────────────────────
# Derivation helpers  (Mode A only)
# ─────────────────────────────────────────────────────────────────────────
def _clean_sku(name):
    return SKU_CLEAN.get(name, name)


def _get_brand(name):
    nl = name.lower()
    if 'sulphur powder' in nl: return 'Sulphur Powder'
    if ('super' in nl or '90' in nl) and ('hariyali' in nl or 'haryali' in nl): return 'Super Hariyali'
    if 'hariyali' in nl or 'haryali' in nl: return 'Hariyali DF'
    if 'profit' in nl: return 'Profit DF'
    if 'ziddi' in nl: return 'Ziddi DF'
    if ('gain' in nl) and ('instabor' not in nl) and ('instacal' not in nl): return 'Gain'
    if 'nutrizin' in nl: return 'NutriZin'
    if 'instabor' in nl or 'insta bor' in nl: return 'InstaBor 150'
    if 'instacal' in nl or 'insta cal' in nl: return 'InstaCal 160'
    if 'instasul' in nl or 'insta sul' in nl: return 'InstaSul'
    return 'Other'


def _get_depot(cn):
    c = cn.lower()
    if 'jaipur' in c:  return 'Jaipur'
    if 'karnal' in c:  return 'Karnal'
    if 'meerut' in c:  return 'Meerut'
    if 'lucknow' in c: return 'Lucknow'
    if 'm.p' in c:     return 'Indore'
    if 'a/20' in c:    return 'Sales (A/20)'
    if 'hisar' in c or 'hissar' in c: return 'Hisar'
    if 'pune' in c:    return 'Pune'
    if 'akola' in c:   return 'Akola'
    if 'bathinda' in c or 'bhatinda' in c: return 'Bathinda'
    if 'gujarat' in c or 'ahmedabad' in c: return 'Ahmedabad'
    if 'raipur' in c:  return 'Raipur'
    return cn


def _extract_party(so_no):
    m = re.search(r'[SA]/\d+/26-27[-\s]*(.*)', str(so_no).strip())
    if m:
        suffix = m.group(1).strip().strip('-').strip()
        parts  = re.split(r'[-\s]+', suffix)
        return ' '.join(p for p in parts if p.upper() not in BRAND_STRIP).strip()
    return ''


def _parse_size(name):
    m = re.search(r'(\d+\.?\d*)\s*(ml|ltr|litre|kg|gm|g\b)', name.lower())
    if not m: return None, None
    val, u = float(m.group(1)), m.group(2)
    if u == 'ml':            return val / 1000, 'ltr'
    if u in ('ltr','litre'): return val,        'ltr'
    if u == 'kg':            return val,        'kg'
    if u in ('gm','g'):      return val / 1000, 'kg'
    return None, None


def _to_kg_ltr(row):
    brand, qty, rate = row['Brand'], row['Balance Qty'], row['Rate']
    if brand in WDG_BRANDS:
        return (round(qty * 1000, 2), 'kg') if rate > MAX_RATE else (round(qty, 2), 'kg')
    # Brand products: Balance Qty is already entered in ltr/kg in the Tally sales order.
    # Just extract the unit from the product name — do NOT multiply by pack size.
    _, unit = _parse_size(row['Product Name'])
    return (round(qty, 3), unit) if unit else (round(qty, 3), 'pcs')


# ─────────────────────────────────────────────────────────────────────────
# Styling helpers  (shared)
# ─────────────────────────────────────────────────────────────────────────
DARK_BLUE  = '1F4E79'
MID_BLUE   = '2E75B6'
LIGHT_BLUE = 'BDD7EE'
PALE_BLUE  = 'DEEAF1'
GREEN_FILL = 'E2EFDA'
ALT_FILL   = 'F2F7FB'
WHITE      = 'FFFFFF'
TOT_ORANGE = 'FCE4D6'
TOT_RED    = 'C00000'
OVERDUE_BG = 'FFE0E0'
YELLOW     = 'FFFF00'


def _fi(hex_): return PatternFill('solid', start_color=hex_)

_thin = Side(style='thin',   color='CCCCCC')
_med  = Side(style='medium', color='888888')


def _b(l=None, r=None, t=None, bo=None):
    return Border(left=l or _thin, right=r or _thin,
                  top=t or _thin,  bottom=bo or _thin)


_STD   = _b()
_BOT   = _b(bo=_med)
_TOPBOT = _b(t=_med, bo=_med)


def _ap(cell, value=None, fill=None, bold=False, color='000000',
        size=9, h='center', wrap=False, bord=None):
    if value is not None: cell.value = value
    cell.fill      = fill or _fi(WHITE)
    cell.font      = Font(name='Arial', bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    cell.border    = bord or _STD


# ─────────────────────────────────────────────────────────────────────────
# Custom errors
# ─────────────────────────────────────────────────────────────────────────
class ProcessingError(Exception):
    """Raised for any user-visible upload / format problem."""
    pass


# ─────────────────────────────────────────────────────────────────────────
# Mode A  —  two raw JSSOReport files
# ─────────────────────────────────────────────────────────────────────────
def _read_jssoreport(f):
    name = getattr(f, 'name', 'uploaded file')
    try:
        f.seek(0)
    except Exception:
        pass
    try:
        df = pd.read_excel(f, sheet_name='JSSOReport', header=0)
    except ValueError as e:
        raise ProcessingError(
            f"'{name}' doesn't contain a sheet named 'JSSOReport'. "
            "For two-file mode please upload the raw system exports unchanged."
        ) from e
    except Exception as e:
        raise ProcessingError(f"Couldn't open '{name}' as an Excel file: {e}") from e
    if 'Product Name' not in df.columns:
        raise ProcessingError(
            f"'{name}' is missing a 'Product Name' column — "
            "it doesn't look like a JSSOReport export."
        )
    try:
        f.seek(0)
    except Exception:
        pass
    return df, name


def _normalize_jssoreport(uploaded_files):
    """
    Read two JSSOReport exports and return a normalized internal DataFrame
    plus a mode-specific summary fragment.

    Internal columns produced:
      Customer Order No., Customer PO Date (datetime), Schedule Date (datetime),
      Customer Name, Party Name, Brand, Depot, Is_Jaishil,
      Product Name, SKU Display, PO Qty, Balance Qty, Rate, Qty_kl, Unit
    """
    f1, f2 = uploaded_files
    df1, name1 = _read_jssoreport(f1)
    df2, name2 = _read_jssoreport(f2)

    has_sulphur1 = df1['Product Name'].str.contains('SULPHUR POWDER', case=False, na=False).any()
    has_sulphur2 = df2['Product Name'].str.contains('SULPHUR POWDER', case=False, na=False).any()
    ambiguous = bool(has_sulphur1 == has_sulphur2)
    if has_sulphur2 and not has_sulphur1:
        new_label, sonepat_label = name2, name1
    else:
        new_label, sonepat_label = name1, name2

    df = pd.concat([df1, df2], ignore_index=True)
    df = df.dropna(subset=['Product Name'])

    for col in ['Customer PO Date', 'Schedule Date']:
        df[col] = pd.to_datetime(df[col], errors='coerce')

    df['Brand']      = df['Product Name'].apply(_get_brand)
    df['Depot']      = df['Customer Name'].apply(_get_depot)
    df['Is_Jaishil'] = df['Customer Name'].str.contains('Jaishil', case=False, na=False)
    df['Party Name'] = df['Customer Order No.'].apply(_extract_party)
    df[['Qty_kl','Unit']] = df.apply(lambda r: pd.Series(_to_kg_ltr(r)), axis=1)
    df['SKU Display'] = df['Product Name'].apply(_clean_sku)

    mode_info = {
        'mode':               'jssoreport',
        'new_file_label':     new_label,
        'sonepat_file_label': sonepat_label,
        'ambiguous_detection': ambiguous,
    }
    return df, mode_info


# ─────────────────────────────────────────────────────────────────────────
# Mode B  —  single pre-built Raw Data file
# ─────────────────────────────────────────────────────────────────────────
def _read_raw_data_sheet(f):
    name = getattr(f, 'name', 'uploaded file')
    try:
        f.seek(0)
    except Exception:
        pass
    try:
        xl = pd.ExcelFile(f)
    except Exception as e:
        raise ProcessingError(f"Couldn't open '{name}' as an Excel file: {e}") from e

    if 'Raw Data' not in xl.sheet_names:
        raise ProcessingError(
            f"'{name}' doesn't have a 'Raw Data' sheet. "
            "For single-file mode, upload the Raw Data export "
            "(the file that already has Brand / Depot / SKU Display / Order Type columns)."
        )
    df = xl.parse('Raw Data', header=0)
    missing = RAW_DATA_REQUIRED_COLS - set(df.columns)
    if missing:
        raise ProcessingError(
            f"'{name}' — Raw Data sheet is missing expected columns: "
            f"{', '.join(sorted(missing))}."
        )
    try:
        f.seek(0)
    except Exception:
        pass
    return df, name


def _normalize_raw_data(uploaded_file):
    """
    Read a single pre-built Raw Data sheet and return a normalized internal
    DataFrame plus a mode-specific summary fragment.

    The Raw Data sheet columns map to internal columns as follows:
      SO / PO No.          → Customer Order No.
      Order Date (str)     → Customer PO Date  (parsed datetime)
      Sched. Date (str)    → Schedule Date     (parsed datetime)
      PO Qty (unit)        → PO Qty
      Pending Qty (unit)   → Balance Qty
      Rate (₹/unit)        → Rate
      Pending (kg or ltr)  → Qty_kl
      Order Type           → Is_Jaishil  (True when 'Inter-depot (Jaishil)')
      All other columns remain as-is.
    """
    df, fname = _read_raw_data_sheet(uploaded_file)

    df = df.rename(columns={
        'SO / PO No.':        'Customer Order No.',
        'PO Qty (unit)':      'PO Qty',
        'Pending Qty (unit)': 'Balance Qty',
        'Rate (₹/unit)':      'Rate',
        'Pending (kg or ltr)':'Qty_kl',
    })

    # Parse date strings — they come out of the formatted sheet as strings like '11-Jun-2026'
    for col_in, col_out in [('Order Date', 'Customer PO Date'), ('Sched. Date', 'Schedule Date')]:
        df[col_out] = pd.to_datetime(df[col_in], format='%d-%b-%Y', errors='coerce')
        if df[col_out].isna().all():
            # Fallback: let pandas try any format
            df[col_out] = pd.to_datetime(df[col_in], errors='coerce')

    df['Is_Jaishil'] = df['Order Type'].str.strip() == 'Inter-depot (Jaishil)'

    # Party Name may or may not be in the file; fill if absent
    if 'Party Name' not in df.columns:
        df['Party Name'] = ''
    df['Party Name'] = df['Party Name'].fillna('')

    # Ensure Rate and Balance Qty exist (needed by External Orders sheet)
    if 'Rate' not in df.columns:
        df['Rate'] = 0.0
    if 'Balance Qty' not in df.columns:
        df['Balance Qty'] = df['Qty_kl']

    mode_info = {
        'mode':        'raw_data',
        'source_file': fname,
    }
    return df, mode_info


# ─────────────────────────────────────────────────────────────────────────
# Workbook builder  (shared by both modes)
# ─────────────────────────────────────────────────────────────────────────
def _build_workbook(df):
    """
    Build and return a formatted openpyxl Workbook from the normalized df.
    Expects the internal column set produced by either _normalize_* function.
    """
    # ── Segment the data ───────────────────────────────────────────────
    jaishil     = df[df['Is_Jaishil']].copy()
    external    = df[~df['Is_Jaishil']].copy()
    sulphur_ext = external[external['Brand'] == 'Sulphur Powder'].copy()
    wdg_ext     = external[external['Brand'] == 'Hariyali DF'].copy()

    # ── Pivot for Production Planning ─────────────────────────────────
    pivot_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    sku_unit   = {}
    for _, r in jaishil.iterrows():
        pivot_data[r['Brand']][r['SKU Display']][r['Depot']] += r['Qty_kl']
        sku_unit[(r['Brand'], r['SKU Display'])] = r['Unit']

    present_depots = [d for d in DEPOT_ORDER if d in jaishil['Depot'].unique()]
    N = len(present_depots)
    COL_BRAND = 1; COL_SKU = 2; COL_UNIT = 3
    COL_DEP_START = 4; COL_TOTAL = 4 + N

    wb = Workbook()

    # ── Sheet 1: Production Planning ───────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Production Planning'
    ws1.freeze_panes = 'C3'

    def block_header(row, title):
        ws1.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=COL_TOTAL)
        _ap(ws1.cell(row=row, column=1), value=title,
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=10, h='left', bord=_BOT)
        ws1.row_dimensions[row].height = 18
        return row + 1

    def col_headers(row):
        ws1.row_dimensions[row].height = 32
        for ci, h in enumerate(
                ['Brand', 'SKU / Pack size', 'Unit'] + present_depots + ['TOTAL\npending'], 1):
            _ap(ws1.cell(row=row, column=ci), value=h,
                fill=_fi(MID_BLUE), bold=True, color=WHITE, size=9, wrap=True, bord=_BOT)
        return row + 1

    def sku_block(start_row, brand_list):
        row = start_row
        depot_grand = {d: 0.0 for d in present_depots}
        grand_total = 0.0
        for brand in brand_list:
            if brand not in pivot_data: continue
            brand_skus = sorted(
                [s for s in pivot_data[brand] if sum(pivot_data[brand][s].values()) > 0])
            if not brand_skus: continue
            depot_brand = {d: 0.0 for d in present_depots}
            brand_total = 0.0
            for i, sku in enumerate(brand_skus):
                ws1.row_dimensions[row].height = 14
                rf = _fi(ALT_FILL) if i % 2 == 0 else _fi(WHITE)
                _ap(ws1.cell(row=row, column=COL_BRAND),
                    value=brand if i == 0 else '', fill=rf, bold=True,
                    color=DARK_BLUE, size=9, h='left')
                _ap(ws1.cell(row=row, column=COL_SKU),
                    value=sku, fill=rf, size=9, h='left')
                _ap(ws1.cell(row=row, column=COL_UNIT),
                    value=sku_unit.get((brand, sku), ''), fill=rf, size=9, color='555555')
                sku_total = 0.0
                for di, depot in enumerate(present_depots):
                    val = pivot_data[brand][sku].get(depot, 0.0)
                    _ap(ws1.cell(row=row, column=COL_DEP_START + di),
                        value=round(val, 3) if val > 0 else None,
                        fill=_fi(GREEN_FILL) if val > 0 else rf, size=9)
                    depot_brand[depot] += val
                    sku_total += val
                _ap(ws1.cell(row=row, column=COL_TOTAL),
                    value=round(sku_total, 3), fill=_fi(TOT_ORANGE),
                    bold=True, color=TOT_RED, size=9)
                brand_total += sku_total
                row += 1
            # Brand subtotal row
            ws1.row_dimensions[row].height = 14
            _ap(ws1.cell(row=row, column=COL_BRAND), value='',
                fill=_fi(LIGHT_BLUE), bord=_TOPBOT)
            _ap(ws1.cell(row=row, column=COL_SKU),
                value=f'  ↳ {brand}  —  TOTAL', fill=_fi(LIGHT_BLUE),
                bold=True, color=DARK_BLUE, size=9, h='left', bord=_TOPBOT)
            _ap(ws1.cell(row=row, column=COL_UNIT), value='',
                fill=_fi(LIGHT_BLUE), bord=_TOPBOT)
            for di, depot in enumerate(present_depots):
                val = depot_brand[depot]
                _ap(ws1.cell(row=row, column=COL_DEP_START + di),
                    value=round(val, 3) if val > 0 else None,
                    fill=_fi(LIGHT_BLUE), bold=True, color=DARK_BLUE, size=9, bord=_TOPBOT)
                depot_grand[depot] += val
            _ap(ws1.cell(row=row, column=COL_TOTAL),
                value=round(brand_total, 3), fill=_fi(LIGHT_BLUE),
                bold=True, color=DARK_BLUE, size=9, bord=_TOPBOT)
            grand_total += brand_total
            row += 1
        # Grand total row
        ws1.row_dimensions[row].height = 16
        _ap(ws1.cell(row=row, column=COL_BRAND), value='TOTAL',
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, bord=_TOPBOT)
        _ap(ws1.cell(row=row, column=COL_SKU), value='All SKUs',
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, h='left', bord=_TOPBOT)
        _ap(ws1.cell(row=row, column=COL_UNIT), value='',
            fill=_fi(DARK_BLUE), bord=_TOPBOT)
        for di, depot in enumerate(present_depots):
            _ap(ws1.cell(row=row, column=COL_DEP_START + di),
                value=round(depot_grand[depot], 3) if depot_grand[depot] > 0 else None,
                fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, bord=_TOPBOT)
        _ap(ws1.cell(row=row, column=COL_TOTAL),
            value=round(grand_total, 3), fill=_fi(DARK_BLUE),
            bold=True, color=YELLOW, size=10, bord=_TOPBOT)
        return row + 2

    r = block_header(1, '▶  SULPHUR / WDG PRODUCTS  (qty in kg)')
    r = col_headers(r)
    r = sku_block(r, WDG_BRAND_ORDER)
    r = block_header(r, '▶  BRAND PRODUCTS  (qty in kg or ltr)')
    r = col_headers(r)
    r = sku_block(r, BRAND_BRAND_ORDER)

    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 44
    ws1.column_dimensions['C'].width = 7
    for di in range(N):
        ws1.column_dimensions[get_column_letter(COL_DEP_START + di)].width = 11
    ws1.column_dimensions[get_column_letter(COL_TOTAL)].width = 13

    # ── Sheet 2: Order Queue (FIFO) ────────────────────────────────────
    ws2 = wb.create_sheet('Order Queue (FIFO)')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 30
    for ci, h in enumerate(
            ['#','Order Date','Sched. Date','SO No.','Depot','Party Name',
             'Brand','SKU / Pack','Pending (kg/ltr)','Unit','Order Total'], 1):
        _ap(ws2.cell(row=1, column=ci), value=h,
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, wrap=True, bord=_BOT)

    order_meta = (
        jaishil
        .sort_values(['Schedule Date','Customer PO Date','Customer Order No.'])
        [['Customer Order No.','Customer PO Date','Schedule Date','Depot','Party Name']]
        .drop_duplicates('Customer Order No.')
        .reset_index(drop=True)
    )
    TODAY = pd.Timestamp(datetime.now().date())
    row2 = 2
    for order_num, (_, om) in enumerate(order_meta.iterrows(), 1):
        so    = om['Customer Order No.']
        lines = jaishil[jaishil['Customer Order No.'] == so].reset_index(drop=True)
        order_total = round(lines['Qty_kl'].sum(), 3)
        od  = om['Customer PO Date'].strftime('%d-%b-%y') if pd.notna(om['Customer PO Date']) else ''
        sd  = om['Schedule Date'].strftime('%d-%b-%y')   if pd.notna(om['Schedule Date'])   else ''
        age = (TODAY - om['Schedule Date']).days          if pd.notna(om['Schedule Date'])   else 0
        overdue   = age > 0
        hdr_fill  = _fi(OVERDUE_BG) if overdue else _fi(PALE_BLUE)
        hdr_color = TOT_RED         if overdue else DARK_BLUE
        age_label = f'  ⚠ {age}d late' if overdue else ''

        for li, (_, line) in enumerate(lines.iterrows()):
            ws2.row_dimensions[row2].height = 14
            rf    = _fi(ALT_FILL) if li % 2 == 0 else _fi(WHITE)
            first = (li == 0)
            _ap(ws2.cell(row=row2,column=1),  value=order_num if first else '',
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9)
            _ap(ws2.cell(row=row2,column=2),  value=od if first else '',
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9)
            _ap(ws2.cell(row=row2,column=3),  value=(sd + age_label) if first else '',
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9)
            _ap(ws2.cell(row=row2,column=4),  value=so if first else '',
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9, h='left')
            _ap(ws2.cell(row=row2,column=5),  value=om['Depot'],
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9)
            _ap(ws2.cell(row=row2,column=6),  value=(om['Party Name'] or '—'),
                fill=hdr_fill if first else rf, bold=first,
                color=hdr_color if first else '000000', size=9, h='left')
            _ap(ws2.cell(row=row2,column=7),  value=line['Brand'],
                fill=rf, size=9, h='left')
            _ap(ws2.cell(row=row2,column=8),  value=line['SKU Display'],
                fill=rf, size=9, h='left')
            _ap(ws2.cell(row=row2,column=9),  value=round(line['Qty_kl'], 3),
                fill=rf, bold=True, size=9)
            _ap(ws2.cell(row=row2,column=10), value=line['Unit'],
                fill=rf, size=9, color='555555')
            _ap(ws2.cell(row=row2,column=11),
                value=order_total if first else '',
                fill=_fi(TOT_ORANGE) if first else rf,
                bold=first, color=TOT_RED if first else '000000', size=9)
            row2 += 1
        for ci in range(1, 12):
            ws2.cell(row=row2, column=ci).border = _b(t=_med)
        row2 += 1

    for ci, w in enumerate([4,11,16,30,13,14,14,44,13,6,13], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: External Orders ───────────────────────────────────────
    ws3 = wb.create_sheet('External Orders')
    ws3.freeze_panes = 'A2'
    ws3.row_dimensions[1].height = 24
    for ci, h in enumerate(
            ['Order Date','Sched. Date','PO / SO No.','Customer Name',
             'Product','Pending (MT)','Pending (kg)'], 1):
        _ap(ws3.cell(row=1, column=ci), value=h,
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, bord=_BOT)

    ext_all = (
        pd.concat([sulphur_ext, wdg_ext])
        .sort_values(['Brand','Customer Name','Customer PO Date'])
        .reset_index(drop=True)
    )
    current_brand = None
    brand_mt = brand_kg = 0.0
    row3 = 2
    for i, (_, r) in enumerate(ext_all.iterrows()):
        if r['Brand'] != current_brand:
            if current_brand is not None:
                for ci, v in enumerate(
                        ['','','',f'TOTAL — {current_brand}','',
                         round(brand_mt, 3), round(brand_kg, 1)], 1):
                    _ap(ws3.cell(row=row3, column=ci), value=v,
                        fill=_fi(TOT_ORANGE), bold=True, color=TOT_RED, size=9,
                        h='left' if ci == 4 else 'center', bord=_TOPBOT)
                row3 += 1
            current_brand = r['Brand']
            brand_mt = brand_kg = 0.0
            ws3.merge_cells(start_row=row3, start_column=1, end_row=row3, end_column=7)
            _ap(ws3.cell(row=row3, column=1), value=f'  {current_brand}',
                fill=_fi(MID_BLUE), bold=True, color=WHITE, size=9, h='left', bord=_BOT)
            row3 += 1
        rf = _fi(ALT_FILL) if i % 2 == 0 else _fi(WHITE)
        od = r['Customer PO Date'].strftime('%d-%b-%y') if pd.notna(r['Customer PO Date']) else ''
        sd = r['Schedule Date'].strftime('%d-%b-%y')   if pd.notna(r['Schedule Date'])   else ''
        for ci, (val, h) in enumerate([
                (od,                        'center'),
                (sd,                        'center'),
                (r['Customer Order No.'],   'left'),
                (r['Customer Name'],        'left'),
                (r['Product Name'],         'left'),
                (round(r['Balance Qty'],3), 'center'),
                (round(r['Qty_kl'],    1),  'center'),
        ], 1):
            _ap(ws3.cell(row=row3, column=ci), value=val, fill=rf, size=9, h=h)
        brand_mt += r['Balance Qty']
        brand_kg += r['Qty_kl']
        row3 += 1
    if current_brand:
        for ci, v in enumerate(
                ['','','',f'TOTAL — {current_brand}','',
                 round(brand_mt, 3), round(brand_kg, 1)], 1):
            _ap(ws3.cell(row=row3, column=ci), value=v,
                fill=_fi(TOT_ORANGE), bold=True, color=TOT_RED, size=9,
                h='left' if ci == 4 else 'center', bord=_TOPBOT)

    for ci, w in enumerate([11,11,26,40,40,13,13], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 4: Raw Data ──────────────────────────────────────────────
    ws4 = wb.create_sheet('Raw Data')
    ws4.freeze_panes = 'A2'
    for ci, h in enumerate([
            'Order Type','Brand','Depot','Customer Name','Party Name',
            'SO / PO No.','Order Date','Sched. Date','Product Name','SKU Display',
            'PO Qty (unit)','Pending Qty (unit)','Rate (₹/unit)',
            'Pending (kg or ltr)','Unit'], 1):
        _ap(ws4.cell(row=1, column=ci), value=h,
            fill=_fi(DARK_BLUE), bold=True, color=WHITE, size=9, bord=_BOT)

    raw = df.copy()
    raw['Order Type']   = raw['Is_Jaishil'].map(
        {True: 'Inter-depot (Jaishil)', False: 'External'})
    raw['Order Date']   = raw['Customer PO Date'].dt.strftime('%d-%b-%Y')
    raw['Sched. Date']  = raw['Schedule Date'].dt.strftime('%d-%b-%Y')
    raw = raw.sort_values(
        ['Order Type','Brand','Depot','Order Date','Customer Order No.'])

    for ri, (_, r) in enumerate(raw.iterrows()):
        er = ri + 2
        ws4.row_dimensions[er].height = 13
        rf = _fi(ALT_FILL) if ri % 2 == 0 else _fi(WHITE)
        for ci, val in enumerate([
                r['Order Type'], r['Brand'], r['Depot'],
                r['Customer Name'], r['Party Name'],
                r['Customer Order No.'], r['Order Date'], r['Sched. Date'],
                r['Product Name'], r['SKU Display'],
                r.get('PO Qty', ''), r['Balance Qty'], r.get('Rate', ''),
                round(r['Qty_kl'], 3), r['Unit'],
        ], 1):
            _ap(ws4.cell(row=er, column=ci), value=val,
                fill=_fi(GREEN_FILL) if ci == 14 else rf,
                bold=(ci == 14), size=9,
                h='left' if ci in [4,5,6,9,10] else 'center')

    for ci, w in enumerate([18,14,13,44,14,28,12,12,44,44,12,14,12,16,6], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ── Save to in-memory buffer ───────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, jaishil, external, sulphur_ext, wdg_ext, present_depots, order_meta


# ─────────────────────────────────────────────────────────────────────────
# Auto-detection helpers
# ─────────────────────────────────────────────────────────────────────────
def _peek_sheets(f):
    """Return the sheet names of an uploaded file without consuming it."""
    try:
        f.seek(0)
    except Exception:
        pass
    try:
        sheets = pd.ExcelFile(f).sheet_names
    except Exception:
        sheets = []
    try:
        f.seek(0)
    except Exception:
        pass
    return sheets


def _detect_mode(uploaded_files):
    """
    Return 'raw_data' or 'jssoreport' based on what was uploaded.
    Raises ProcessingError with a clear message if neither pattern matches.
    """
    n = len(uploaded_files)
    if n == 1:
        sheets = _peek_sheets(uploaded_files[0])
        name   = getattr(uploaded_files[0], 'name', 'the uploaded file')
        if 'Raw Data' in sheets:
            return 'raw_data'
        if 'JSSOReport' in sheets:
            raise ProcessingError(
                f"You uploaded one file ('{name}') that contains a 'JSSOReport' sheet. "
                "For two-file JSSOReport mode please upload **both** exports — "
                "the plant file and the Sonepat file."
            )
        raise ProcessingError(
            f"'{name}' has neither a 'Raw Data' sheet nor a 'JSSOReport' sheet. "
            "Please upload either:\n"
            "• The pre-built Raw Data export (one file, 'Raw Data' sheet), or\n"
            "• Both raw JSSOReport exports (two files, each with a 'JSSOReport' sheet)."
        )
    if n == 2:
        sheets0 = _peek_sheets(uploaded_files[0])
        sheets1 = _peek_sheets(uploaded_files[1])
        name0   = getattr(uploaded_files[0], 'name', 'File 1')
        name1   = getattr(uploaded_files[1], 'name', 'File 2')
        # Reject if someone uploads two Raw Data files
        if 'Raw Data' in sheets0 and 'Raw Data' in sheets1:
            raise ProcessingError(
                "Both files look like pre-built Raw Data exports. "
                "For single-file mode, upload just **one** Raw Data file."
            )
        # If one is Raw Data and one is JSSOReport — ambiguous
        if 'Raw Data' in sheets0 or 'Raw Data' in sheets1:
            raise ProcessingError(
                "You've uploaded a mix of file types. For two-file mode both "
                "files must be raw JSSOReport exports. "
                "For single-file mode upload only the Raw Data file."
            )
        if 'JSSOReport' not in sheets0:
            raise ProcessingError(
                f"'{name0}' doesn't contain a 'JSSOReport' sheet. "
                "Please upload unmodified raw system exports."
            )
        if 'JSSOReport' not in sheets1:
            raise ProcessingError(
                f"'{name1}' doesn't contain a 'JSSOReport' sheet. "
                "Please upload unmodified raw system exports."
            )
        return 'jssoreport'
    raise ProcessingError(
        f"Please upload 1 or 2 files (you uploaded {n})."
    )


# ─────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────
def process(uploaded_files):
    """
    Auto-detect the input mode, normalize the data, build the workbook,
    and return (buffer, summary).

    Parameters
    ----------
    uploaded_files : list of file-like objects (1 or 2 items)

    Returns
    -------
    buffer  : io.BytesIO  — the finished .xlsx workbook
    summary : dict        — counts and metadata for the UI
    """
    uploaded_files = list(uploaded_files)
    mode = _detect_mode(uploaded_files)

    if mode == 'jssoreport':
        df, mode_info = _normalize_jssoreport(uploaded_files)
    else:
        df, mode_info = _normalize_raw_data(uploaded_files[0])

    # Exclude internal ledger account
    excluded_a20 = df[df['Depot'] == 'Sales (A/20)']
    excluded_a20_count = int(len(excluded_a20))
    df = df[df['Depot'] != 'Sales (A/20)'].copy()

    buf, jaishil, external, sulphur_ext, wdg_ext, present_depots, order_meta = \
        _build_workbook(df)

    TODAY = pd.Timestamp(datetime.now().date())
    overdue_orders = order_meta[
        order_meta['Schedule Date'] < TODAY
    ]['Customer Order No.'].tolist()

    summary = {
        **mode_info,
        'jaishil_orders':       int(jaishil['Customer Order No.'].nunique()),
        'external_order_lines': int(len(sulphur_ext) + len(wdg_ext)),
        'overdue_count':        len(overdue_orders),
        'overdue_orders':       overdue_orders,
        'excluded_a20_rows':    excluded_a20_count,
        'total_rows_processed': int(len(df)),
        'depots_present':       present_depots,
    }
    return buf, summary
